import warnings
warnings.filterwarnings("ignore")
import os
import sys
sys.path.append(os.path.abspath(os.path.dirname(__file__)))
base_dir = os.path.abspath(os.path.dirname(__file__))

import numpy as np
from openpyxl import load_workbook

import math
import itertools
from multiprocessing import Pool, cpu_count

from make_PTDF import PTDF
import opf_es_carbon


def value_function(x_input, u_t, TG_num, RG_num, TG_carbon, TG_offer, TG_maxG, TG_minG,
                   RG_offer, RG_P, RG_cap_t, D_P_t,
                   branch_max, PTDF, A_TG, A_RG, A_D):
    return opf_es_carbon.opf_carbon(
        u_t, TG_carbon, TG_offer,
        TG_maxG * x_input[: TG_num],
        TG_minG * x_input[: TG_num],
        RG_offer, RG_P * x_input[TG_num : TG_num + RG_num],
        RG_cap_t, D_P_t * x_input[TG_num + RG_num:],
        branch_max, PTDF, A_TG, A_RG, A_D
    )['carbon_t']

def value_total(u_t, TG_carbon, TG_offer, TG_maxG, TG_minG,
                RG_offer, RG_P, RG_cap_t, D_P_t,
                branch_max, PTDF, A_TG, A_RG, A_D):
    return opf_es_carbon.opf_carbon(u_t, TG_carbon, TG_offer, TG_maxG, TG_minG,
                      RG_offer, RG_P, RG_cap_t, D_P_t,
                      branch_max, PTDF, A_TG, A_RG, A_D)['carbon_t']

def calculate_shapley_for_agent(args):
    (i, n, value_function, shapley_values,
     u_t, TG_num, RG_num, TG_carbon, TG_offer, TG_maxG, TG_minG,
     RG_offer, RG_P, RG_cap_t, D_P_t,
     branch_max, PTDF, A_TG, A_RG, A_D) = args

    marginal_contribution_sum = 0
    for subset in itertools.chain.from_iterable(
            itertools.combinations(list(range(n))[:i] + list(range(n))[i + 1:], r) for r in range(n)
    ):
        # print(subset)
        # 创建两个子集，一个是包含i的子集，一个是没有i的子集
        subset_with_i = np.zeros(n)  # 包含i的子集
        subset_without_i = np.zeros(n)  # 不包含i的子集
        # # 遍历当前子集的所有元素
        for j in subset:
            subset_without_i[j] = 1  # 不包含i的子集赋值
            subset_with_i[j] = 1  # 包含i的子集赋值
        # 添加玩家i到包含i的子集
        subset_with_i[i] = 1
        marginal_contribution = value_function(
            subset_with_i, u_t, TG_num, RG_num, TG_carbon, TG_offer, TG_maxG, TG_minG,
            RG_offer, RG_P, RG_cap_t, D_P_t,
            branch_max, PTDF, A_TG, A_RG, A_D
        ) - value_function(
            subset_without_i, u_t, TG_num, RG_num, TG_carbon, TG_offer, TG_maxG, TG_minG,
            RG_offer, RG_P, RG_cap_t, D_P_t,
            branch_max, PTDF, A_TG, A_RG, A_D
        )
        # 子集的权重 | S |!(n - | S | -1)! / n!
        weight = math.factorial(len(subset)) * math.factorial(n - len(subset) - 1) / math.factorial(n)
        marginal_contribution_sum += weight * marginal_contribution
        print(marginal_contribution_sum)
    shapley_values[i] = marginal_contribution_sum
    return marginal_contribution_sum

def shapley_value_parallel(n, value_function, u_t, TG_carbon, TG_offer, TG_maxG, TG_minG,
                           RG_offer, RG_P, RG_cap_t, D_P_t,
                           branch_max, PTDF, A_TG, A_RG, A_D):
    # max_processes = int(os.getenv('SLURM_CPUS_PER_TASK'))
    max_processes = min(8, cpu_count())  # 默认为cpu_count()
    shapley_values = np.zeros(n)

    with Pool(max_processes) as pool:
        args = [(i, n, value_function, shapley_values,
                 u_t, TG_num, RG_num, TG_carbon, TG_offer, TG_maxG, TG_minG,
                 RG_offer, RG_P, RG_cap_t, D_P_t,
                 branch_max, PTDF, A_TG, A_RG, A_D) for i in range(n)]
        results = pool.map(calculate_shapley_for_agent, args)

    for i, result in enumerate(results):
        shapley_values[i] = result

    return shapley_values

if __name__ == "__main__":
    MAIN_CASE_ID = 3
    pkl_file_path = os.path.join(base_dir, 'data', f'case_example_dict_{MAIN_CASE_ID}.pkl')
    output_dir = os.path.join(base_dir, 'Shapley_value_results')
    os.makedirs(output_dir, exist_ok=True)

    PTDF_result = PTDF(pkl_file_path)

    PTDF = PTDF_result['PTDF']
    branch_max = PTDF_result['branch_max']
    A_TG = PTDF_result['A_TG']
    A_RG = PTDF_result['A_RG']
    A_D = PTDF_result['A_D']
    case_example = PTDF_result['case_example']

    TG_num = int(case_example['TG_num'])
    RG_num = int(case_example['RG_num'])
    D_num = int(case_example['D_num'])
    ES_num = int(case_example['ES_num'])

    u = case_example['u_TG'].astype(int)
    TG_carbon = case_example['TG_carbon'].astype(float)
    TG_offer = case_example['TG_offer'].astype(float)
    TG_maxG = case_example['TG_maxG'].astype(float)
    TG_minG = case_example['TG_minG'].astype(float)
    RG_offer = case_example['RG_offer'].astype(float)
    RG_P = case_example['RG_P'].astype(float)
    RG_cap = case_example['RG_cap'].astype(float)
    D_P = case_example['D_P'].astype(float)

    n = TG_num + RG_num + D_num

    senaid = MAIN_CASE_ID
    t = 0
    u_t = u [t, :]
    D_P_t = D_P[t, :]
    RG_cap_t = RG_cap[t, :]

    carbon_sharing = shapley_value_parallel(
        n, value_function, u_t, TG_carbon, TG_offer, TG_maxG, TG_minG,
        RG_offer, RG_P, RG_cap_t, D_P_t,
        branch_max, PTDF, A_TG, A_RG, A_D
    )

    carbon_total = sum(carbon_sharing)

    carbon_origin = value_total(
        u_t, TG_carbon, TG_offer, TG_maxG, TG_minG,
        RG_offer, RG_P, RG_cap_t, D_P_t,
        branch_max, PTDF, A_TG, A_RG, A_D
    )

    excel_path = os.path.join(output_dir, "Shapley_value_results.xlsx")
    wb = load_workbook(excel_path)
    ws = wb.active

    row_data = [t] + carbon_sharing.tolist() + [carbon_total, carbon_origin]

    for col, val in enumerate(row_data, start=1):
        ws.cell(row=t + 2, column=col).value = val

    wb.save(excel_path)

    print(f"=== shapley_value_main_t_{t} 处理完成 ===")
    if carbon_total == carbon_origin:
        print(f"Carbon emission obligation allocation for t=0:\n{carbon_sharing}\n, Total carbon emission:{carbon_total}")
    else:
        print(f"Carbon emission obligation allocation for t=0:\n{carbon_sharing}\n, Total carbon emission:{carbon_total}. Origin carbon emission:{carbon_origin}")
